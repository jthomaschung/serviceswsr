"""
Jimmy John's WSR Export Bot - with Download Audit & Retry
Automates downloading WSR (Weekly Sales Report) exports from Jimmy John's portal
"""

import io
import os
import re
import time
import logging
import zipfile
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Set, Tuple
import json
from playwright.sync_api import sync_playwright, Page, Download
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure logging with UTF-8 encoding to handle special characters
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('jj_wsr_bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Full expected store list — update if fleet changes
EXPECTED_STORES = {
    522, 746, 799, 833, 838, 877, 930, 965, 1002, 1018, 1019, 1061, 1111,
    1127, 1206, 1261, 1307, 1337, 1342, 1355, 1440, 1441, 1554, 1556, 1562,
    1635, 1694, 1695, 1696, 1762, 1779, 1789, 1955, 1956, 1957, 2006, 2021,
    2176, 2178, 2180, 2391, 2500, 2501, 2502, 2503, 2504, 2601, 2682, 2683,
    2711, 2712, 2749, 2807, 2808, 2811, 2812, 2821, 2873, 2874, 2876, 2883,
    2884, 3029, 3030, 3187, 3260, 3391, 3612, 3613, 3635, 3686, 3972, 4018,
    4022, 4024, 4105, 4330, 4358, 4586,
}

MAX_RETRY_ATTEMPTS = 2  # How many times to retry missing stores after initial run


class JimmyJohnsWSRBot:
    """Bot for downloading WSR reports from Jimmy John's Macromatix portal"""

    def __init__(self):
        self.start_url = "https://prod-services.jimmyjohns.com/pages/aspx/dashboard/"
        self.email = os.getenv('JJ_EMAIL')
        self.password = os.getenv('JJ_PASSWORD')
        self.download_dir = Path(os.getenv('DOWNLOAD_DIR', './downloads'))
        self.processed_dir = Path(os.getenv('PROCESSED_DIR', './processed'))
        self.download_dir.mkdir(exist_ok=True)
        self.processed_dir.mkdir(exist_ok=True)
        self.database_url = os.getenv('DATABASE_URL')
        self.downloaded_files = []

    # =========================================================================
    # AUDIT HELPERS
    # =========================================================================

    @staticmethod
    def _validate_wsr_file_content(file_data: bytes, filename: str) -> bool:
        """
        Validate the Weekly Sales sheet inside an extracted WSR Excel file.

        Catches the Macromatix-overload signature where the WSR line items
        render correctly but the Expected Deposits section at the bottom of
        the sheet ships back blank. The specific check is: row 9 must contain
        at least one valid date in the AM columns (4, 6, 8, 10, 12, 14, 16).
        A blank date row is the root cause of parse_weekly_sales_file()
        returning zero records for an otherwise-present store.

        Returns True if the file looks complete enough to parse both WSR
        line items and Expected Deposits. Returns False on any failure
        (missing sheet, blank dates, unreadable file).
        """
        lower = filename.lower()
        try:
            if lower.endswith('.xlsx'):
                wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
                if "Weekly Sales" not in wb.sheetnames:
                    return False
                ws = wb["Weekly Sales"]
                # Row 9, AM columns at 4, 6, 8, 10, 12, 14, 16 (1-indexed)
                dates_found = sum(
                    1 for c in range(4, 18, 2)
                    if ws.cell(9, c).value not in (None, "")
                )
                return dates_found >= 7
            elif lower.endswith('.xls'):
                df = pd.read_excel(
                    io.BytesIO(file_data),
                    sheet_name="Weekly Sales",
                    header=None,
                    engine="xlrd",
                )
                # Row 9 in sheet = index 8 in 0-indexed df
                if df.shape[0] <= 8:
                    return False
                dates_found = 0
                for c0_1based in range(4, 18, 2):
                    c0 = c0_1based - 1
                    if c0 < df.shape[1] and pd.notna(df.iat[8, c0]):
                        dates_found += 1
                return dates_found >= 7
            else:
                return False
        except Exception as e:
            logger.debug(f"  Content validation error for {filename}: {e}")
            return False

    def get_downloaded_stores_for_week(
        self,
        week_str: str,
        validate_content: bool = False,
    ) -> Tuple[Set[int], Set[int]]:
        """
        Scan the processed directory for ZIP files matching week_str.

        Returns a tuple of (valid_stores, invalid_stores):
          • valid_stores    — store numbers whose file passed all checks
          • invalid_stores  — store numbers whose file was found but had
                              corrupt/blank content (only populated when
                              validate_content=True)

        week_str format matches the ZIP filename, e.g. '03-03-2026'.
        """
        valid_stores: Set[int] = set()
        invalid_stores: Set[int] = set()

        for zip_path in self.processed_dir.glob(f"WSR_Export_{week_str}_*.zip"):
            try:
                with zipfile.ZipFile(zip_path, 'r') as zf:
                    for name in zf.namelist():
                        # Files are named like "#3030 03-03-26.xls"
                        m = re.match(r'#(\d+)\s', name)
                        if not m:
                            continue
                        store_num = int(m.group(1))

                        if not validate_content:
                            valid_stores.add(store_num)
                            continue

                        # Content validation — read file bytes from the ZIP
                        # and confirm the Weekly Sales sheet is populated
                        try:
                            with zf.open(name) as inner:
                                data = inner.read()
                            if self._validate_wsr_file_content(data, name):
                                valid_stores.add(store_num)
                            else:
                                invalid_stores.add(store_num)
                                logger.warning(
                                    f"  Content validation FAILED for store "
                                    f"{store_num} in {zip_path.name} "
                                    f"(blank dates or missing sheet)"
                                )
                        except Exception as e:
                            invalid_stores.add(store_num)
                            logger.warning(
                                f"  Could not validate {name} in {zip_path.name}: {e}"
                            )
            except Exception as e:
                logger.warning(f"Could not inspect ZIP {zip_path.name}: {e}")

        # A store is only "valid" if it has at least one valid copy. If a
        # store appears in both sets (bad ZIP + good ZIP from an earlier
        # retry), treat it as valid and drop it from invalid.
        invalid_stores -= valid_stores
        return valid_stores, invalid_stores

    def audit_week_downloads(
        self,
        week_str: str,
        validate_content: bool = True,
    ) -> Set[int]:
        """
        Compare downloaded stores against EXPECTED_STORES.

        With validate_content=True (default), the audit also opens each
        store's Excel file and verifies the Weekly Sales sheet has the
        date row populated. Stores with corrupt files are returned in
        the "missing" set so the existing retry loop re-downloads them.

        Returns the set of missing OR corrupt store numbers (empty = all good).
        """
        valid, invalid = self.get_downloaded_stores_for_week(
            week_str, validate_content=validate_content
        )
        not_downloaded = EXPECTED_STORES - valid - invalid
        missing = (EXPECTED_STORES - valid)  # union of not_downloaded + invalid

        logger.info(f"\n{'─'*60}")
        logger.info(f"DOWNLOAD AUDIT — week {week_str}")
        logger.info(f"  Expected : {len(EXPECTED_STORES)} stores")
        logger.info(f"  Valid    : {len(valid)} stores")
        if validate_content and invalid:
            logger.warning(
                f"  ⚠️  CORRUPT content ({len(invalid)}): {sorted(invalid)}"
            )
        if not_downloaded:
            logger.warning(
                f"  ⚠️  NOT downloaded ({len(not_downloaded)}): {sorted(not_downloaded)}"
            )
        if not missing:
            logger.info(f"  ✓ All stores accounted for with valid content")

        return missing

    # =========================================================================
    # LOGIN / NAVIGATION  (unchanged from original)
    # =========================================================================

    def login(self, page: Page) -> bool:
        """Handle login if needed"""
        try:
            logger.info("Navigating to Jimmy John's portal...")
            page.goto(self.start_url, wait_until='domcontentloaded', timeout=30000)
            page.wait_for_timeout(2000)

            if "dashboard" in page.url.lower() and page.locator('text="MY DASHBOARD"').count() > 0:
                logger.info("Already logged in!")
                return True

            if page.locator('input[type="email"], input[type="text"]').count() > 0:
                logger.info("Login required, entering credentials...")
                email_input = page.locator('input[type="email"], input[type="text"]').first
                email_input.fill(self.email)

                if page.locator('button:has-text("NEXT")').count() > 0:
                    page.locator('button:has-text("NEXT")').click()
                    page.wait_for_timeout(2000)

                password_input = page.locator('input[type="password"]').first
                password_input.fill(self.password)

                for selector in ['button:has-text("SIGN IN")', 'button:has-text("Sign In")',
                                  'button:has-text("Login")', 'button[type="submit"]']:
                    if page.locator(selector).count() > 0:
                        page.locator(selector).first.click(timeout=60000, no_wait_after=True)
                        break

                # Portal can be slow to redirect after sign-in
                try:
                    page.wait_for_load_state('networkidle', timeout=60000)
                except Exception:
                    pass  # Fall through to URL check below
                page.wait_for_timeout(5000)

            if "dashboard" in page.url.lower() or page.locator('text="MY DASHBOARD"').count() > 0:
                logger.info("Successfully on dashboard!")
                return True
            else:
                logger.error(f"Login failed. Current URL: {page.url}")
                page.screenshot(path='login_failed.png')
                return False

        except Exception as e:
            logger.error(f"Login process failed: {e}")
            page.screenshot(path='login_error.png')
            return False

    def navigate_to_wsr_export(self, page: Page) -> bool:
        """Navigate to WSR Export page"""
        try:
            logger.info("Looking for Sales Reports link...")
            sales_reports = page.locator('text="Sales Reports"')
            if sales_reports.count() > 0:
                sales_reports.click()
                # Use 'load' instead of 'networkidle' — the JJ portal is an SPA with
                # persistent background XHR traffic that prevents networkidle from ever firing.
                try:
                    page.wait_for_load_state('load', timeout=15000)
                except Exception:
                    pass  # domcontentloaded/load already fired per logs; safe to continue
                page.wait_for_timeout(2000)

                for selector in ['text="WSR EXPORT"', 'text="WSR Export"',
                                  'a:has-text("WSR")', '*:has-text("WSR EXPORT")']:
                    if page.locator(selector).count() > 0:
                        page.locator(selector).first.click()
                        try:
                            page.wait_for_load_state('load', timeout=10000)
                        except Exception:
                            pass
                        page.wait_for_timeout(2000)

                        if page.locator('text="Select Reporting Week Ending Date"').count() > 0:
                            logger.info("Successfully on WSR Export page")
                            return True
                        break

                if page.locator('text="Select Reporting Week Ending Date"').count() == 0:
                    logger.error("Could not find WSR Export page elements")
                    page.screenshot(path='wsr_navigation_failed.png')
                    return False
            else:
                logger.error("Could not find Sales Reports link")
                page.screenshot(path='sales_reports_not_found.png')
                return False

            return True

        except Exception as e:
            logger.error(f"Failed to navigate to WSR Export: {e}")
            page.screenshot(path='navigation_error.png')
            return False

    def select_reporting_week(self, page: Page, week_offset: int = 0) -> str:
        """Select reporting week (0 = most recent, 1 = previous week, etc.)"""
        try:
            week_dropdown = page.locator('text="Select Reporting Week Ending Date"').locator('xpath=following-sibling::*').first
            week_dropdown.click()
            page.wait_for_timeout(1000)

            week_options = page.locator('[role="option"]').all()
            if not week_options:
                week_options = page.locator('.dropdown-item').all()

            if week_offset < len(week_options):
                selected_week = week_options[week_offset].text_content()
                week_options[week_offset].click()
                logger.info(f"Selected week: {selected_week}")
                return selected_week
            else:
                logger.error(f"Week offset {week_offset} out of range")
                return None

        except Exception as e:
            logger.error(f"Failed to select week: {e}")
            return None

    def open_stores_dropdown(self, page: Page) -> bool:
        """Find and open the Stores dropdown"""
        try:
            candidates = page.locator('div[class*="select"], div[class*="dropdown"], mat-select, ng-select').all()
            logger.info(f"Found {len(candidates)} select-like elements on page")

            clicked = False
            for idx, candidate in enumerate(candidates):
                try:
                    if candidate.is_visible():
                        candidate.click()
                        page.wait_for_timeout(500)
                        if page.locator('input[type="checkbox"]').count() > 1:
                            logger.info(f"Stores dropdown opened via candidate index {idx}")
                            clicked = True
                            break
                        else:
                            page.keyboard.press('Escape')
                            page.wait_for_timeout(300)
                except Exception:
                    continue

            if not clicked:
                page.locator('text="Stores"').first.click()
                page.wait_for_timeout(500)
                if page.locator('input[type="checkbox"]').count() <= 1:
                    logger.error("Could not open Stores dropdown")
                    page.screenshot(path='stores_dropdown_failed.png')
                    return False

            try:
                page.wait_for_function(
                    "document.querySelectorAll('input[type=\"checkbox\"]').length > 10",
                    timeout=10000
                )
            except Exception:
                pass

            num_checkboxes = page.locator('input[type="checkbox"]').count()
            logger.info(f"Stores dropdown loaded with {num_checkboxes} checkboxes")
            return num_checkboxes > 1

        except Exception as e:
            logger.error(f"open_stores_dropdown failed: {e}")
            return False

    def get_all_stores(self, page: Page) -> int:
        """Get count of all available stores"""
        try:
            if not self.open_stores_dropdown(page):
                return 79

            num_checkboxes = page.locator('input[type="checkbox"]').count()
            page.keyboard.press('Escape')
            page.wait_for_timeout(500)

            num_stores = num_checkboxes - 1 if num_checkboxes > 0 else 0
            if num_stores < 70:
                logger.warning(f"Only {num_stores} stores detected, defaulting to 79")
                return 79

            return num_stores

        except Exception as e:
            logger.error(f"Failed to get store count: {e}")
            return 79

    def select_store_batch(self, page: Page, batch_start: int, batch_size: int = 5, total_stores: int = 80) -> int:
        """Select a batch of stores by checkbox index"""
        try:
            batch_end = min(batch_start + batch_size, total_stores)
            logger.info(f"Selecting stores {batch_start + 1} to {batch_end} of {total_stores}...")

            if not self.open_stores_dropdown(page):
                return 0

            all_checkboxes = page.locator('input[type="checkbox"]').all()
            if not all_checkboxes:
                return 0

            select_all = all_checkboxes[0]
            if select_all.is_checked():
                select_all.click()
                page.wait_for_timeout(500)

            selected_count = 0
            for i in range(batch_start + 1, min(batch_end + 1, len(all_checkboxes))):
                try:
                    if not all_checkboxes[i].is_checked():
                        all_checkboxes[i].click()
                        selected_count += 1
                        page.wait_for_timeout(100)
                except Exception as e:
                    logger.warning(f"Failed to select store at index {i}: {e}")

            page.keyboard.press('Escape')
            page.wait_for_timeout(500)
            logger.info(f"Selected {selected_count} stores in this batch")
            return selected_count

        except Exception as e:
            logger.error(f"Failed to select store batch: {e}")
            return 0

    def select_specific_stores_by_number(self, page: Page, store_numbers: Set[int]) -> int:
        """
        Select specific stores by matching their store number in the checkbox label.
        Used for retry runs targeting only the missing stores.
        """
        try:
            logger.info(f"Selecting {len(store_numbers)} specific stores: {sorted(store_numbers)}")

            if not self.open_stores_dropdown(page):
                return 0

            all_checkboxes = page.locator('input[type="checkbox"]').all()
            if not all_checkboxes:
                return 0

            # Uncheck Select All if checked
            if all_checkboxes[0].is_checked():
                all_checkboxes[0].click()
                page.wait_for_timeout(500)

            selected_count = 0
            # Each checkbox should have an associated label — try to read it
            # Labels are typically in a sibling or parent element
            checkbox_locators = page.locator('input[type="checkbox"]').all()

            for i, cb in enumerate(checkbox_locators[1:], start=1):  # skip Select All
                try:
                    # Get the text near this checkbox (parent li or label)
                    label_text = page.evaluate(
                        """(el) => {
                            const parent = el.closest('li') || el.parentElement;
                            return parent ? parent.textContent.trim() : '';
                        }""",
                        cb.element_handle()
                    )

                    # Extract store number from label text (e.g. "#3030" or "3030")
                    m = re.search(r'#?(\d{3,5})', label_text)
                    if m:
                        store_num = int(m.group(1))
                        if store_num in store_numbers:
                            if not cb.is_checked():
                                cb.click()
                                selected_count += 1
                                page.wait_for_timeout(100)
                                logger.info(f"  ✓ Selected store {store_num}")
                except Exception as e:
                    logger.warning(f"Could not read label for checkbox {i}: {e}")
                    continue

            page.keyboard.press('Escape')
            page.wait_for_timeout(500)
            logger.info(f"Selected {selected_count} of {len(store_numbers)} targeted stores")
            return selected_count

        except Exception as e:
            logger.error(f"select_specific_stores_by_number failed: {e}")
            return 0

    def download_wsr_export(self, page: Page, week: str, batch_num: int) -> Optional[str]:
        """Download the WSR export file"""
        logger.info(f"Starting download for batch {batch_num}...")

        export_button = page.locator('button:has-text("EXPORT")')
        if export_button.count() == 0:
            logger.error("Could not find EXPORT button")
            return None

        # Wrap expect_download separately so a TimeoutError returns None
        # cleanly instead of bubbling up past the batch retry loop
        try:
            with page.expect_download(timeout=120000) as download_info:
                export_button.click()
                logger.info("Clicked EXPORT — waiting up to 2 minutes...")
        except Exception as e:
            logger.error(f"Download timed out or failed waiting for file: {e}")
            return None

        try:
            download = download_info.value
            suggested_filename = download.suggested_filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            week_str = week.replace('/', '-') if week else "unknown"
            extension = Path(suggested_filename).suffix if suggested_filename else '.zip'
            filename = f"WSR_Export_{week_str}_Batch{batch_num}_{timestamp}{extension}"

            save_path = self.download_dir / filename
            download.save_as(save_path)

            processed_path = self.processed_dir / filename
            save_path.rename(processed_path)

            file_size = processed_path.stat().st_size
            logger.info(f"Saved: {filename} ({file_size:,} bytes)")

            if file_size < 1000:
                logger.warning("File too small — may be corrupt")
                return None

            self.downloaded_files.append(processed_path)
            return str(processed_path)

        except Exception as e:
            logger.error(f"Failed to save downloaded file: {e}")
            return None

    # =========================================================================
    # RETRY LOGIC
    # =========================================================================

    def retry_missing_stores(self, page: Page, week_offset: int,
                              selected_week: str, missing_stores: Set[int],
                              week_str: str) -> Set[int]:
        """
        Attempt to download only the missing stores.
        Downloads them in batches of 15 and returns any still-missing stores after.
        """
        logger.info(f"\n{'='*60}")
        logger.info(f"RETRY — downloading {len(missing_stores)} missing stores")
        logger.info(f"{'='*60}")

        missing_list = sorted(missing_stores)
        batch_size = 5
        num_batches = (len(missing_list) + batch_size - 1) // batch_size

        for batch_num in range(num_batches):
            batch_stores = set(missing_list[batch_num * batch_size:(batch_num + 1) * batch_size])
            logger.info(f"\n--- Retry Batch {batch_num + 1} of {num_batches}: {sorted(batch_stores)} ---")

            # Reload page and re-select week
            for _reload_try in range(3):
                try:
                    page.reload()
                    page.wait_for_load_state('networkidle', timeout=30000)
                    break
                except Exception as _reload_err:
                    if _reload_try == 2:
                        raise
                    logger.warning(
                        f"  Page reload timed out ({_reload_try + 1}/3), "
                        f"waiting 5 s before retry..."
                    )
                    page.wait_for_timeout(5000)
            page.wait_for_timeout(3000)
            self.select_reporting_week(page, week_offset)
            page.wait_for_timeout(10000)

            num_selected = self.select_specific_stores_by_number(page, batch_stores)

            if num_selected > 0:
                filepath = self.download_wsr_export(page, selected_week, f"Retry{batch_num + 1}")
                if filepath:
                    logger.info(f"✓ Retry batch {batch_num + 1} downloaded")
                else:
                    logger.warning(f"Retry batch {batch_num + 1} download failed")

                if batch_num < num_batches - 1:
                    time.sleep(10)
            else:
                logger.warning(f"Could not select stores for retry batch {batch_num + 1}")

        # Re-audit after retry
        still_missing = self.audit_week_downloads(week_str)
        return still_missing

    # =========================================================================
    # MAIN RUN
    # =========================================================================

    def refetch_stores(
        self,
        store_numbers: Set[int],
        week_offset: int = 0,
        max_attempts: int = MAX_RETRY_ATTEMPTS,
    ) -> Tuple[Set[int], List[str]]:
        """
        Public entry point for targeted re-downloads. Called by the parser
        when audits show stores missing from the database tables — spins up
        a fresh Playwright session, logs in, and runs retry_missing_stores
        until all requested stores have valid files on disk or max_attempts
        is exhausted.

        Returns (still_missing, downloaded_filepaths):
          • still_missing — stores we could not recover after all attempts
          • downloaded_filepaths — ZIPs created during this refetch
            (the caller can re-extract just these to avoid rescanning the
            whole processed/ directory)
        """
        if not store_numbers:
            return set(), []

        logger.info(f"\n{'='*60}")
        logger.info(
            f"REFETCH — parser requested re-download of "
            f"{len(store_numbers)} stores: {sorted(store_numbers)}"
        )
        logger.info(f"{'='*60}")

        files_before = set(self.processed_dir.glob("WSR_Export_*.zip"))
        missing = set(store_numbers)

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=['--disable-blink-features=AutomationControlled'],
            )
            context = browser.new_context(
                accept_downloads=True,
                viewport={'width': 1920, 'height': 1080},
            )
            page = context.new_page()
            page.on("console", lambda msg: logger.debug(f"Browser: {msg.text}"))

            try:
                if not self.login(page):
                    raise Exception("Login failed during refetch")

                if not self.navigate_to_wsr_export(page):
                    raise Exception("Failed to navigate to WSR Export during refetch")

                selected_week = self.select_reporting_week(page, week_offset)
                if not selected_week:
                    raise Exception(f"Failed to select week offset {week_offset} during refetch")

                week_str = selected_week.strip().replace('/', '-')

                # ── Retry loop (same pattern as run()) ─────────────────
                attempt = 0
                while missing and attempt < max_attempts:
                    attempt += 1
                    logger.info(f"\n--- Refetch attempt {attempt}/{max_attempts} ---")
                    missing = self.retry_missing_stores(
                        page, week_offset, selected_week, missing, week_str
                    )
            finally:
                time.sleep(2)
                browser.close()

        # Figure out which ZIPs are new (downloaded during this refetch)
        files_after = set(self.processed_dir.glob("WSR_Export_*.zip"))
        new_files = [str(p) for p in sorted(files_after - files_before)]

        if missing:
            logger.error(
                f"❌ Refetch incomplete: {len(missing)} stores still missing "
                f"after {max_attempts} attempts: {sorted(missing)}"
            )
        else:
            logger.info(
                f"✓ Refetch complete: recovered all {len(store_numbers)} stores "
                f"({len(new_files)} new ZIP(s))"
            )

        return missing, new_files

    def run(self, weeks_to_download: int = 1):
        """Main execution — downloads all stores with audit and retry"""
        logger.info(f"\n{'='*60}")
        logger.info(f"Starting Jimmy John's WSR Export Bot")
        logger.info(f"{'='*60}\n")

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=['--disable-blink-features=AutomationControlled']
            )
            context = browser.new_context(
                accept_downloads=True,
                viewport={'width': 1920, 'height': 1080}
            )
            page = context.new_page()
            page.on("console", lambda msg: logger.debug(f"Browser: {msg.text}"))

            if not self.login(page):
                raise Exception("Login failed")

            if not self.navigate_to_wsr_export(page):
                raise Exception("Failed to navigate to WSR Export")

            # Track overall audit results
            audit_summary = {}

            for week_offset in range(weeks_to_download):
                logger.info(f"\n{'='*50}")
                logger.info(f"Processing Week {week_offset + 1} of {weeks_to_download}")
                logger.info(f"{'='*50}")

                selected_week = self.select_reporting_week(page, week_offset)
                if not selected_week:
                    logger.error(f"Failed to select week {week_offset}")
                    continue

                # Derive the week_str used in ZIP filenames (e.g. "03-03-2026")
                week_str = selected_week.strip().replace('/', '-') if selected_week else "unknown"

                total_stores = self.get_all_stores(page)
                if total_stores == 0:
                    logger.error("No stores found")
                    continue

                batch_size = 5
                num_batches = (total_stores + batch_size - 1) // batch_size
                logger.info(f"Total stores: {total_stores} | Batches: {num_batches}")

                # ── Initial download pass ──────────────────────────────────
                BATCH_MAX_ATTEMPTS = 3

                for batch_num in range(num_batches):
                    batch_start = batch_num * batch_size
                    logger.info(f"\n--- Batch {batch_num + 1} of {num_batches} ---")

                    batch_success = False

                    for attempt in range(1, BATCH_MAX_ATTEMPTS + 1):
                        if attempt > 1:
                            logger.warning(f"  Batch {batch_num + 1} attempt {attempt}/{BATCH_MAX_ATTEMPTS}...")

                        # Reload and re-select week before every attempt
                        # (always reload so state is clean, skip only on very first attempt of first batch)
                        if batch_num > 0 or attempt > 1:
                            _reload_ok = False
                            for _reload_try in range(3):
                                try:
                                    page.reload(timeout=30000, wait_until='domcontentloaded')
                                    page.wait_for_load_state('networkidle', timeout=30000)
                                    _reload_ok = True
                                    break
                                except Exception as _reload_err:
                                    logger.warning(
                                        f"  Page reload attempt ({_reload_try + 1}/3) failed: "
                                        f"{type(_reload_err).__name__} — waiting 5 s before retry..."
                                    )
                                    page.wait_for_timeout(5000)

                            if not _reload_ok:
                                # Frame is detached or page context is broken — navigate fresh
                                logger.warning(
                                    "  All reload attempts failed (frame may be detached). "
                                    "Navigating fresh to WSR Export page..."
                                )
                                try:
                                    page.goto(
                                        self.start_url,
                                        wait_until='domcontentloaded',
                                        timeout=45000,
                                    )
                                    page.wait_for_load_state('networkidle', timeout=30000)
                                except Exception as _nav_err:
                                    logger.error(f"  Fresh navigation to dashboard failed: {_nav_err}")
                                    raise

                                if not self.navigate_to_wsr_export(page):
                                    raise RuntimeError(
                                        "Could not re-navigate to WSR Export after frame detach"
                                    )
                                logger.info("  Successfully restored WSR Export page context.")

                            page.wait_for_timeout(2000)
                            self.select_reporting_week(page, week_offset)
                            page.wait_for_timeout(10000)

                        num_selected = self.select_store_batch(page, batch_start, batch_size, total_stores)

                        if num_selected == 0:
                            logger.warning(f"  No stores selected on attempt {attempt}")
                            continue

                        filepath = self.download_wsr_export(page, selected_week, batch_num + 1)

                        if filepath:
                            batch_success = True
                            break
                        else:
                            logger.warning(f"  Download failed on attempt {attempt}")

                    if not batch_success:
                        raise Exception(
                            f"Batch {batch_num + 1} failed after {BATCH_MAX_ATTEMPTS} attempts "
                            f"(stores {batch_start + 1}–{min(batch_start + batch_size, total_stores)}). "
                            f"Aborting run."
                        )

                    if batch_num < num_batches - 1:
                        time.sleep(10)

                # ── Audit ─────────────────────────────────────────────────
                missing = self.audit_week_downloads(week_str)

                # ── Retry loop ────────────────────────────────────────────
                attempt = 0
                while missing and attempt < MAX_RETRY_ATTEMPTS:
                    attempt += 1
                    logger.info(f"\n{'='*50}")
                    logger.info(f"RETRY ATTEMPT {attempt}/{MAX_RETRY_ATTEMPTS} for week {week_str}")
                    logger.info(f"{'='*50}")

                    missing = self.retry_missing_stores(
                        page, week_offset, selected_week, missing, week_str
                    )

                if missing:
                    logger.error(
                        f"❌ Week {week_str}: {len(missing)} stores still missing after "
                        f"{MAX_RETRY_ATTEMPTS} retries: {sorted(missing)}"
                    )
                    audit_summary[week_str] = {"status": "INCOMPLETE", "missing": sorted(missing)}
                else:
                    logger.info(f"✓ Week {week_str}: All stores downloaded successfully")
                    audit_summary[week_str] = {"status": "COMPLETE", "missing": []}

            time.sleep(5)
            browser.close()

        # ── Final summary ─────────────────────────────────────────────────
        logger.info(f"\n{'='*60}")
        logger.info("BOT EXECUTION COMPLETE — AUDIT SUMMARY")
        logger.info(f"{'='*60}")
        for week, result in audit_summary.items():
            status = result["status"]
            missing = result["missing"]
            if missing:
                logger.error(f"  {week}: {status} — missing {len(missing)} stores: {missing}")
            else:
                logger.info(f"  {week}: {status} ✓")
        logger.info(f"Total files downloaded: {len(self.downloaded_files)}")

        # Exit with error code if any week is incomplete (fails the GitHub Action)
        if any(r["status"] == "INCOMPLETE" for r in audit_summary.values()):
            raise Exception("Download audit failed — one or more weeks are incomplete. Check logs.")


def main():
    bot = JimmyJohnsWSRBot()
    weeks_to_download = int(os.getenv('WEEKS_TO_DOWNLOAD', 1))
    bot.run(weeks_to_download)


if __name__ == "__main__":
    main()
